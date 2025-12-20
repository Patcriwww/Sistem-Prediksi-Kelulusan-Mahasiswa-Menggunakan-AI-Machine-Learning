import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================
# UTIL
# =========================
def _df(rows):
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows).copy()
    for c in ["probability", "ipk", "presensi", "sks_lulus", "mengulang"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

def _summary(df):
    if df.empty:
        return dict(total=0, low=0, med=0, high=0, lowp=0, medp=0, highp=0)

    total = len(df)
    low = (df["risk"] == "Low Risk").sum()
    med = (df["risk"] == "Medium Risk").sum()
    high = (df["risk"] == "High Risk").sum()

    pct = lambda x: round((x / total) * 100, 1) if total else 0
    return {
        "total": total,
        "low": low, "lowp": pct(low),
        "med": med, "medp": pct(med),
        "high": high, "highp": pct(high),
    }

def _latest_high_risk(df):
    if df.empty:
        return df
    latest = df.sort_values("timestamp").drop_duplicates("nim", keep="last")
    return latest[latest["risk"] == "High Risk"].sort_values("probability", ascending=False)

# =========================
# EXPORT EXCEL (STYLED)
# =========================
def export_excel(rows, path):
    df = _df(rows)
    summary = _summary(df)
    high = _latest_high_risk(df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ---- SUMMARY SHEET ----
        summary_df = pd.DataFrame([{
            "Total Prediksi": summary["total"],
            "Low Risk (%)": f'{summary["low"]} ({summary["lowp"]}%)',
            "Medium Risk (%)": f'{summary["med"]} ({summary["medp"]}%)',
            "High Risk (%)": f'{summary["high"]} ({summary["highp"]}%)',
        }])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        ws = writer.book["Summary"]
        ws.freeze_panes = "A2"

        header_fill = PatternFill("solid", fgColor="E8F0FE")
        header_font = Font(bold=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # ---- HIGH RISK SHEET ----
        if not high.empty:
            cols = [
                "nama_mahasiswa", "nim", "kelas", "probability",
                "ipk", "presensi", "sks_lulus", "recommendation"
            ]
            high[cols].to_excel(writer, sheet_name="HighRisk", index=False)
        else:
            pd.DataFrame(columns=["nama_mahasiswa","nim","risk"]).to_excel(
                writer, sheet_name="HighRisk", index=False
            )

        ws = writer.book["HighRisk"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FDECEA")

        # ---- ALL DATA ----
        if not df.empty:
            df.to_excel(writer, sheet_name="AllData", index=False)
            ws = writer.book["AllData"]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True)

        # ---- AUTO WIDTH ----
        for sheet in writer.book.worksheets:
            for col in sheet.columns:
                length = max(len(str(c.value)) if c.value else 0 for c in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = min(length + 4, 30)

# =========================
# EXPORT PDF (PROFESSIONAL)
# =========================
def export_pdf(rows, path):
    df = _df(rows)
    s = _summary(df)
    high = _latest_high_risk(df)

    c = canvas.Canvas(path, pagesize=A4)
    width, height = A4
    y = height - 60

    # Colors
    blue = HexColor("#2563EB")
    gray = HexColor("#6B7280")
    red = HexColor("#DC2626")

    # Title
    c.setFont("Helvetica-Bold", 18)
    c.setFillColor(blue)
    c.drawString(50, y, "Laporan Akademik Berbasis AI")
    y -= 26

    c.setFont("Helvetica", 11)
    c.setFillColor(gray)
    c.drawString(50, y, "Prediksi Kelulusan Mahasiswa & Intervensi Rekomendasi (ARaaS)")
    y -= 24

    # Summary Box
    c.setFillColorRGB(0.95, 0.96, 1)
    c.roundRect(50, y-80, width-100, 80, 10, fill=1, stroke=0)
    c.setFillColorRGB(0, 0, 0)

    c.setFont("Helvetica-Bold", 12)
    c.drawString(60, y-20, "Ringkasan")
    c.setFont("Helvetica", 10)

    c.drawString(60, y-40, f"Total Prediksi : {s['total']}")
    c.drawString(240, y-40, f"Low Risk : {s['low']} ({s['lowp']}%)")
    c.drawString(60, y-60, f"Medium Risk : {s['med']} ({s['medp']}%)")
    c.drawString(240, y-60, f"High Risk : {s['high']} ({s['highp']}%)")

    y -= 110

    # High Risk Table
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Mahasiswa High Risk (Prediksi Terbaru)")
    y -= 18

    c.setFont("Helvetica", 9)
    if high.empty:
        c.drawString(50, y, "Tidak ada mahasiswa High Risk.")
        y -= 14
    else:
        for _, r in high.iterrows():
            line = (
                f"{r['nama_mahasiswa']} ({r['nim']}) | "
                f"Kelas {r['kelas']} | "
                f"{r['probability']}%"
            )
            c.drawString(50, y, line[:110])
            y -= 12
            if y < 60:
                c.showPage()
                y = height - 60
                c.setFont("Helvetica", 9)

    # Footer
    c.setFont("Helvetica-Oblique", 8)
    c.setFillColor(gray)
    c.drawString(
        50, 40,
        f"Dihasilkan otomatis pada {datetime.now().strftime('%d %B %Y %H:%M')} • Sistem Prediksi Kelulusan Mahasiswa"
    )

    c.save()
